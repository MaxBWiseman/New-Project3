import csv
import os
import requests
from bs4 import BeautifulSoup
from collections import Counter
from datetime import datetime
import itertools
import threading
import sys
import time
import re
import math
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil import parser
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi
from geopy.distance import geodesic
from dotenv import load_dotenv
from flask import Flask, send_from_directory

app = Flask(__name__)

load_dotenv()

#Directory to save Excel and CSV files
UPLOAD_FOLDER = 'data_visuals'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/')
def index():
    files = os.listdir(app.config['UPLOAD_FOLDER'])
    file_links = ''.join([f'<li><a href="/download/{file}">{file}</a></li>' for file in files])
    return f'''
        <h1 style='text-align:center;'>Welcome to the Event Exporter!</h1>
        <ul style='text-align:center;'>{file_links}</ul>
        
    '''

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


def start_flask_server():
    print('Flask server started. Press CNTRL-C to quit exporting.')
    app.run(debug=False)

class Spinner:
    def __init__(self, message='Loading...'):
        self.message = message
        self.spinner = itertools.cycle(['|', '/', '-', '\\'])
        self.stop_running = threading.Event()
# This sets up the message and spinner, and creates a stop_running event that will be used to stop the spinner.
        
    def start(self):
        threading.Thread(target=self._spin).start()
# This starts a new thread that will run the _spin method.
    
    def _spin(self):
        while not self.stop_running.is_set():
# This method will run as long as the stop_running event is not set.
            sys.stdout.write(f'\r{self.message} {next(self.spinner)}')
# \r is a carriage return, which moves the cursor to the beginning of the line.
# This allows the spinner to overwrite itself on the same line, next(self.spinner) gets the next character in the spinner.
            sys.stdout.flush()
# Immediately flushes to the console to show the spinner.
            time.sleep(0.1)
            sys.stdout.write('\b')
# \b is a backspace, which moves the cursor back one character so a new character can be written.
    
    def stop(self):
        self.stop_running.set()
# Thread is stopped
        sys.stdout.write('\r' + ' ' * (len(self.message) + 2) + '\r')
# Clean up the spinner by overwriting it with spaces and moving the cursor back to the beginning of the line.
        sys.stdout.flush()


# Hashtable to cache recently searched events
cache = {}

uri = os.getenv('MONGO_URI')
# Create a new client and connect to the server
client = MongoClient(uri, server_api=ServerApi('1'))
# Send a ping to confirm a successful connection
try:
    client.admin.command('ping')
    print("Pinged your deployment. You successfully connected to MongoDB!")
except Exception as e:
    print(e)
    
db = client['Event_Hoarder']
collection = db['Event_Data']

def check_and_delete_old_events():
    current_date = datetime.now().date()
    events = collection.find({})
    
    for event in events:
        unique_id = event.get('url', 'N/A')
        if unique_id == 'N/A':
            continue
        
        start_date = event.get('event_date_time', 'N/A')
        
        try:
            checked_start_date = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S').date()
        except ValueError:
            collection.delete_one({'url': unique_id})
            continue
        
        if current_date > checked_start_date:
            collection.delete_one({'url': unique_id})

check_and_delete_old_events()

def save_to_mongodb(collection, search_key, events):
   
    for event in events:
        unique_id = event.get('url', 'N/A')
        if unique_id == 'N/A':
            continue
        
        event_data = {
            'search_key': search_key,
            'url': unique_id,
            'name': event.get('name', 'N/A'),
            'location': event.get('location', 'N/A'),
            'event_date_time': event.get('event_date_time', 'N/A'),
            'show_date_time': event.get('show_date_time', 'N/A'),
            'summary': event.get('summary', 'N/A'),
            'event_price': event.get('event_price', 'N/A')
        }
        
        collection.update_one({'url': unique_id}, {'$set': event_data}, upsert=True)


def parsed_scraped_date(date_time):
    if 'No date and time available' in date_time or not date_time.strip():
        return 'N/A'
    # If the date_time is not available, return 'N/A'
    
    replacements = {
        'Monday': '',
        'Mon': '',
        'Tuesday': '',
        'Tue': '',
        'Wednesday': '',
        'Wed': '',
        'Thursday': '',
        'Thu': '',
        'Friday': '',
        'Fri': '',
        'Saturday': '',
        'Sat': '',
        'Sunday': '',
        'Sun': '',
        'Starts on': '',
        'GMT': '',
        'GMT+1': '',
        'January': 'Jan',
        'February': 'Feb',
        'March': 'Mar',
        'April': 'Apr',
        'May': 'May',
        'June': 'Jun',
        'July': 'Jul',
        'August': 'Aug',
        'September': 'Sep',
        'October': 'Oct',
        'November': 'Nov',
        'December': 'Dec',
        'pm': 'PM',
        'am': 'AM',
        '·': '',
        ' - ': ' ',
        '+1': '',
    }

    for key, value in replacements.items():
        date_time = date_time.replace(key, value)
    # Replace the long names of the days and months with their abbreviations
    # Also replace other unwanted strings

    # Remove any extra spaces and commas
    date_time = ' '.join(date_time.split()).replace(',', '')
    
    # For cases where users enter date ranges, only take the first date
    date_time_parts = date_time.split(' ')
    # Only take the first 5 parts of the date_time
    if len(date_time_parts) > 3:
        date_time = ' '.join(date_time_parts[:3])
    # If the length of the date_time is less than 3, return the date_time as is
    # If the length of the date_time is greater than 3, only take the first 5 parts of the date_time
    # parts example: ['2024-10-19', '16:30:00']

    try:
        # Use dateutil.parser to parse the date
        dt = parser.parse(date_time, fuzzy=True)
    except ValueError:
        raise ValueError(f"Date format not recognized: {date_time}")
    # parser.parse will try to parse the date and time from the string
    # fuzzy=True allows for more flexibility in the date format

    # Format the datetime object into the desired format
    formatted_date = dt.strftime('%Y-%m-%d %H:%M:%S')

    return formatted_date
    
    

def scrape_eventbrite_events(location, day, product, page_number, start_date, end_date):
    url = f'https://www.eventbrite.com/d/united-kingdom--{location}/events--{day}/{product}/?page={page_number}&start_date={start_date}&end_date={end_date}'
    
    page = requests.get(url)
    
    soup = BeautifulSoup(page.content, 'html.parser')
    
    events = soup.find_all('a', class_='event-card-link')
    
    event_data = []
    tags_counter = Counter()
    seen_urls = set()
    
    for event in events:
        event_url = event['href']
        if event_url in seen_urls:
            continue
        seen_urls.add(event_url)
        
        event_info = {
            'name': event.get('aria-label', '').replace('View', '').strip(),
            'url': event_url
        }
        
        page_detail = requests.get(event_url)
        page_detail_soup = BeautifulSoup(page_detail.content, 'html.parser')

        price_div = page_detail_soup.find('div', class_="conversion-bar__panel-info")
        event_price = price_div.get_text(strip=True) if price_div else 'Free'
        
        location_div = page_detail_soup.find('div', class_='location-info__address')
        if location_div:
            event_location = location_div.get_text(separator=' ', strip=True)
            # Code with help from Co-Pilot
            event_location= re.sub(r'Show map$', '', event_location)
            # End of Co-Pilot code, this code removes the 'Show map' text from the location
            # even if its preceeded by another word, the issue was "United KingdomShow map"
        else:
            location_div = page_detail_soup.find('div', class_='location-info__address-text')
            if location_div:
                event_location = location_div.get_text(separator=' ', strip=True)

        summary_div = page_detail_soup.find('div', class_='eds-text--left')
        event_summary = ''
        if summary_div:
            p_elements = summary_div.find_all('p')
            summary_texts = [p_element.get_text(separator=' ',strip=True) for p_element in p_elements]
            event_summary = ' '.join(summary_texts)
        elif not summary_div:
            summary = page_detail_soup.find('p', class_='summary')
            event_summary = summary.get_text(strip=True) if summary else 'No summary available'

        date_time = page_detail_soup.find('span', class_='date-info__full-datetime')
        event_date_time = date_time.get_text(strip=True) if date_time else 'No date and time available'

        date_parsed = parsed_scraped_date(event_date_time)
        
        tags = page_detail_soup.find_all('a', class_='tags-link')
        for tag in tags:
            tags_counter[tag.get_text(strip=True)] += 1

        event_info.update({
            'location': event_location,
            'show_date_time': event_date_time, # More clearer user version of the date
            'event_date_time': date_parsed,
            'summary': event_summary,
            'event_price': event_price
        })

        event_data.append(event_info)
    
    return event_data, tags_counter


# FIXME: Scrape top events seems to only scrape about half the events on the top results page for the given area, must look into this
def scrape_eventbrite_top_events(country, day, location, category_slug, page_number, start_date, end_date):
    url = f'https://www.eventbrite.co.uk/d/{country}--{location}/{category_slug}--events--{day}/?page={page_number}&start_date={start_date}&end_date={end_date}'
    
    page = requests.get(url)
    
    soup = BeautifulSoup(page.content, 'html.parser')
    
    events = soup.find_all('a', class_='event-card-link')
    
    event_data = []
    tags_counter = Counter()
    seen_urls = set()
    
    for event in events:
        event_url = event['href']
        if event_url in seen_urls:
            continue
        seen_urls.add(event_url)
        
        event_info = {
            'name': event.get('aria-label', '').replace('View', '').strip(),
            'url': event_url
        }
        
        page_detail = requests.get(event_url)
        page_detail_soup = BeautifulSoup(page_detail.content, 'html.parser')

        price_div = page_detail_soup.find('span', class_="eds-text-bm eds-text-weight--heavy")
        event_price = price_div.get_text(strip=True) if price_div else 'Free'

        location_div = page_detail_soup.find('div', class_='location-info__address')
        if location_div:
            event_location = location_div.get_text(separator=' ', strip=True)
            # Code with help from Co-Pilot
            event_location= re.sub(r'Show map$', '', event_location)
            # End of Co-Pilot code, this code removes the 'Show map' text from the location
            # even if its preceeded by another word, the issue was "United KingdomShow map"
        else:
            location_div = page_detail_soup.find('div', class_='location-info__address-text')
            if location_div:
                event_location = location_div.get_text(separator=' ', strip=True)
        
        summary_div = page_detail_soup.find('div', class_='eds-text--left')
        event_summary = ''
        if summary_div:
            p_elements = summary_div.find_all('p')
            summary_texts = [p_element.get_text(separator=' ', strip=True) for p_element in p_elements]
            event_summary = ' '.join(summary_texts)
        elif not summary_div:
            summary = page_detail_soup.find('p', class_='summary')
            event_summary = summary.get_text(strip=True) if summary else 'No summary available'
                

        date_time = page_detail_soup.find('span', class_='date-info__full-datetime')
        event_date_time = date_time.get_text(strip=True) if date_time else 'No date and time available'

        date_parsed = parsed_scraped_date(event_date_time)
        
        tags = page_detail_soup.find_all('a', class_='tags-link')
        for tag in tags:
            tags_counter[tag.get_text(strip=True)] += 1

        event_info.update({
            'location': event_location,
            'show_date_time': event_date_time,
            'event_date_time': date_parsed,
            'summary': event_summary,
            'event_price': event_price
        })

        event_data.append(event_info)
    
    return event_data, tags_counter, len(event_data)


def scrape_eventbrite_top_events_no_category(location, country):
    url = f'https://www.eventbrite.co.uk/d/{country}--{location}/events/'
    
    page = requests.get(url)
    
    soup = BeautifulSoup(page.content, 'html.parser')
    
    events = soup.find_all('a', class_='event-card-link')
    
    event_data = []
    tags_counter = Counter()
    seen_urls = set()
    
    for event in events:
        event_url = event['href']
        if event_url in seen_urls:
            continue
        seen_urls.add(event_url)
        
        event_info = {
            'name': event.get('aria-label', '').replace('View', '').strip(),
            'url': event_url
        }
        
        page_detail = requests.get(event_url)
        page_detail_soup = BeautifulSoup(page_detail.content, 'html.parser')
        
        location_div = page_detail_soup.find('div', class_='location-info__address')
        if location_div:
            event_location = location_div.get_text(separator=' ', strip=True)
            # Code with help from Co-Pilot
            event_location= re.sub(r'Show map$', '', event_location)
            # End of Co-Pilot code, this code removes the 'Show map' text from the location
            # even if its preceeded by another word, the issue was "United KingdomShow map"
        else:
            location_div = page_detail_soup.find('div', class_='location-info__address-text')
            if location_div:
                event_location = location_div.get_text(separator=' ', strip=True)
        
        price_div = page_detail_soup.find('span', class_="eds-text-bm eds-text-weight--heavy")
        event_price = price_div.get_text(strip=True) if price_div else 'Free'

        summary_div = page_detail_soup.find('div', class_='eds-text--left')
        event_summary = ''
        if summary_div:
            p_elements = summary_div.find_all('p')
            summary_texts = [p_element.get_text(separator=' ', strip=True) for p_element in p_elements]
            event_summary = ' '.join(summary_texts)
        elif not summary_div:
            summary = page_detail_soup.find('p', class_='summary')
            event_summary = summary.get_text(strip=True) if summary else 'No summary available'


        date_time = page_detail_soup.find('span', class_='date-info__full-datetime')
        event_date_time = date_time.get_text(strip=True) if date_time else 'No date and time available'

        date_parsed = parsed_scraped_date(event_date_time)
        
        event_info.update({
            'location': event_location,
            'show_date_time': event_date_time,
            'event_date_time': date_parsed,
            'summary': event_summary,
            'event_price': event_price
        })

        event_data.append(event_info)
    return event_data

def save_to_csv(events):
    directory = 'data_visuals'
    file_name = os.path.join(directory, 'collected_events.csv')
    file_exists = os.path.exists(file_name)
    
    fields = ['name', 'location', 'show_date_time', 'event_price', 'summary', 'url']

    with open(file_name, 'a', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=fields)

        if not file_exists:
            writer.writeheader()

        for event in events:
            filtered_event = {field: event[field] for field in fields if field in event}
            # this will only include the fields that are in the fields list above, so that
            # fields that only have a programmatic purpose are not included in the CSV
            # it works by creating a new dictionary with only the fields that are in the fields list
            writer.writerow(filtered_event)

    print(f"\n-------------------------------------\nEvents saved to {file_name}, download/view from the main menu.\n-------------------------------------")
    return

def save_to_excel(events, filename='data_visuals/events_data.xlsx'):
    
    filename = check_file_unique(filename)
    
    workbook = openpyxl.Workbook()
    # Create a new Excel workbook
    sheet = workbook.active
    # active means the first sheet in the workbook
    sheet.title = 'Events Data'
    # Set the title of the sheet to 'Events Data'
    
    headers = ['Event Name', 'Date', 'Location', 'Price', 'Summary', 'URL']
    column_widths = [71, 58, 111, 12, 81, 140]
    # Set the column headers and their widths
    
    for col_num, (header, width) in enumerate(zip(headers, column_widths), 1):
        # zip() pairs each element from headers with the corresponding element from column_widths,
        # this creates an iterator of tuples where each tuple contains a header and its width
        # enumerate() numerates each iteration of the tuples list, starting at 1, giving the column number
        # Result: [(1, ('Event Name', 71)), (2, ('Date', 58)), (3, ('Location', 111)) etc.]
        col_letter = get_column_letter(col_num)
        # get_column_letter() is a built in function from openpyxl that returns the letter of the specified column, example: 1 -> 'A', 2 -> 'B'
        sheet[f'{col_letter}1'] = header
        # Will set the headers stated above in the first row of the sheet iterating through columns A, B, C, D, E, F 
        sheet.column_dimensions[col_letter].width = width
        # column_dimensions is a dictionary that stores the width of each column, the width is set to the width in the column_widths list
        
    for row_num, event in enumerate(events, 2):
        sheet[f'A{row_num}'] = event.get('name', 'N/A')
        sheet[f'B{row_num}'] = event.get('show_date_time', 'N/A')
        sheet[f'C{row_num}'] = event.get('location', 'N/A')
        sheet[f'D{row_num}'] = event.get('event_price', 'N/A')
        sheet[f'E{row_num}'] = event.get('summary', 'N/A')
        sheet[f'F{row_num}'] = event.get('url', 'N/A')
    
    workbook.save(filename)
    print(f"\n-------------------------------------\nEvents saved to {filename}, download/view from the main menu.\n-------------------------------------")

def collection_menu():
    while True:
        print("\nChoose an option to manipulate events or print to CSV:")
        print("1. View all searched events")
        print("2. View recent searches")
        print("3. something")
        print("4. Main Menu")
        print("#. Clear Database")
        choice = input("Enter your choice: ").strip()

        if choice == '1':
            view_all_events()
        elif choice == '2':
            search_events_in_collection()
        elif choice == '3':
            something()
        elif choice == '4':
            print("Going back to the main menu.")
            main()
        elif choice == '#':
            collection.delete_many({})
            print('-------------------------------------\nDatabase cleared\n-------------------------------------.')
            main()
        else:
            print("Invalid choice. Please try again.")

def get_unique_search_keys():
    pipeline = [
        {'$group': {'_id': '$search_key'}},
    ]
    unique_search_keys = list(collection.aggregate(pipeline))
    return [key['_id'] for key in unique_search_keys]
# This function will return a list of unique search keys from the mongodb collection
# The pipeline will group the documents by the search_key field and return only unique values

# This function was built with help from Co-Pilot when asked how could i extract the price from the event_price field
# Its suggestion was the re module. As i needed to remove other strings and currency symbols from the event_price field
def extract_price(price_str):
    # Use regular expression to find the first occurrence of a number in the string
    match = re.search(r'\d+(\.\d+)?', price_str)
    # \d+(\.\d+)? means match one or more digits followed by an optional decimal point and one or more digits
    if match:
        return float(match.group())
    return 0.0
"""
The extract_price function uses a regular expression to find the first occurrence of a number in the event_price string.
If a number is found, it is converted to a float and returned. If no number is found, 0.0 is returned.
The sort_events function uses the extract_price function to extract the numeric part of the event_price before sorting. 
"""

def check_file_unique(image_path):
    directory, filename = os.path.split(image_path)
    # Example: 'data_visuals/event_count_per_day.png' -> ('data_visuals', 'event_count_per_day.png')
    base, ext = os.path.splitext(filename)
    # Example: 'event_count_per_day.png' -> ('event_count_per_day', '.png')
    counter = 1
    
    while os.path.exists(image_path):
        image_path = os.path.join(directory, f'{base}_{counter}{ext}')
        # While theres already a file with the same name, add a counter to the filename
        counter += 1
    
    return image_path

def get_coordinates(location, api_key):
    url = f'https://maps.googleapis.com/maps/api/geocode/json?address={location}&key={api_key}'
    # Construct the URL for the Google Maps Geocoding API
    response = requests.get(url)
    # Send a GET request to the URL
    if response.status_code == 200:
        data = response.json()
    # If the response is successful, turn the response into a JSON object
        if data['status'] == 'OK':
    # Check if the status in the JSON object is 'OK'
            location = data['results'][0]['geometry']['location']
            return (location['lat'], location['lng'])
    # Extract the latitude and longitude from the JSON object and return them as a tuple
    return None

def find_closest_events(user_location, events, api_key):
    user_coordinates = get_coordinates(user_location, api_key)
    # Get the coordinates of the user's location
    if not user_coordinates:
        print("User location could not be geocoded.")
        return []

    def distance_to_user(event):
        event_coordinates = get_coordinates(event['location'], api_key)
        # Get the coordinates of the event's location
        if event_coordinates:
            return geodesic(user_coordinates, event_coordinates).miles
        # The geopy library's geodesic function calculates the distance between two points on the Earth's surface
        # using the geodesic distance, which is more accurate than the haversine formula, as it accounts for the earths ellipsoidal shape,
        # it returns the distance in miles. Sourch - https://www.askpython.com/python/examples/find-distance-between-two-geo-locations
        return float('inf')
        # If fails return infinity as this is a easy way to sort the events with no coordinates to the end of the list
        
    events_with_coordinates = [event for event in events if event.get('location')]
    # Filter out events with no location
    events_with_coordinates.sort(key=distance_to_user)
    # Sort the events by using the distance_to_user function as the key
    
    return events_with_coordinates[::-1]
    # Return the sorted events in reverse order

def compare_events(events):
    if len(events) < 2:
        print('Not enough events to compare.')
        return
    
    print('\nWhat would you like to compare?')
    print('1. Average price of events')
    print('2. Median price of events')
    print('3. Event count per day')
    print('4. Event count per month')
    print('5. Event price distribution')
    print('6. Event dates over time')
    print('8. Compare organizers')
    print('9. Main Menu')
    choice = input('Enter your choice: ').strip()
    
    if choice == '1':
        price = [extract_price(event['event_price']) for event in events if event.get('event_price', '').lower() not in ['sold out', 'free', 'donation']]
        # Extract the price from the event_price field for each event, not including 'sold out', 'free', and 'donation' events
        result = sum(price) / len(price) if price else 0
        # Calculate the average price of the events, sum the prices and divide by the number of prices
        floored_result = math.floor(result)
        # Round down the result to the nearest whole number
        print(f'\nThe average price of events is: £{floored_result}')
    elif choice == '2':
        price = [extract_price(event['event_price']) for event in events if event.get('event_price', '').lower() not in ['sold out', 'free', 'donation']]
        # Extract the price from the event_price field for each event, not including 'sold out', 'free', and 'donation' events
        if price:
            price.sort()
        # Sort the list of prices in ascending order
            n = len(price)
        # Get length of list
            mid = n // 2
        # Find the middle index
            if n % 2 == 0:
                result = (price[mid - 1] + price[mid]) / 2
        # If the list length is even, the median is the average of the two middle elements
            else:
                result = price[mid]
        # If the list length is odd, the median is the middle element
        else:
            result = 0
        # If empty list, return 0
        print(f'\nThe median price of events is: £{result}')
    elif choice == '3':
        event_days = [datetime.strptime(event['event_date_time'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%d') for event in events]
        # List comprehension to extract the day and year from the event_date_time field for each event, the date is expected to be in the format 'YYYY-MM-DD HH:MM:SS' and is parsed to a datetime object.
        # The second argument formats the datetime object to 'YYYY-DD' to be used to group the events by day
        days_counts = Counter(event_days)
        # Use the Counter class to count the occurrences of each day, works by setting a dictionary key with each collected day and incrementing the value each time the day is found
        days, counts = zip(*sorted(days_counts.items()))
        # Take the dictionary days_counts and retrieve its items as a list of tuples with items(), then sort the tuples by the keys (days) sorted(), take the sorted list of tuples and unpacks them into two lists, one for keys and one for values zip(*).
        # the splat operator * allowed zip() to take the list of tuples and unpack them into two seperate lists, without it zip() would return one list of tuples
        plt.bar(days, counts)
        # Create a bar chart with the days on the x-axis and the counts on the y-axis
        plt.title('Events By Day')
        plt.xlabel
        plt.ylabel('Number of Events')
        plt.xticks(rotation=45)
        # Rotate the x-axis labels by 45 degrees for better readability
        plt.tight_layout()
        # Automatic padding
        image_path = 'data_visuals/event_count_per_day.png'
        image_path = check_file_unique(image_path)
        plt.savefig(image_path)
        # Save the plot as an image
        plt.close()
        # I decided to use the matplotlib library to create various data visualizations
        print(f'\n-------------------------------------\nEvent count per day saved as {image_path}, download/view from the main menu.\n-------------------------------------')
    elif choice == '4':
        event_months = [datetime.strptime(event['event_date_time'], '%Y-%m-%d %H:%M:%S').strftime('%Y-%m') for event in events]
        # List comprehension to extract the month and year from the event_date_time field for each event, the date is expected to be in the format 'YYYY-MM-DD HH:MM:SS' and is parsed to a datetime object.
        # The second argument formats the datetime object to 'YYYY-MM' to be used to group the events by month
        months_counts = Counter(event_months)
        # Use the Counter class to count the occurrences of each month, works by setting a dictionary key with each collected month and incrementing the value each time the month is found
        months, counts = zip(*sorted(months_counts.items()))
        # Take the dictionary months_counts and retrieve its items as a list of tuples with items(), then sort the tuples by the keys (months) sorted(), take the sorted list of tuples and unpacks them into two lists, one for keys and one for values zip(*).
        # the splat operator * allowed zip() to take the list of tuples and unpack them into two seperate lists, without it zip() would return one list of tuples
        plt.bar(months, counts)
        # Create a bar chart with the months on the x-axis and the counts on the y-axis
        plt.title('Events By Month')
        plt.xlabel
        plt.ylabel('Number of Events')
        plt.xticks(rotation=45)
        # Rotate the x-axis labels by 45 degrees for better readability
        plt.tight_layout()
        # Automatic padding
        image_path = 'data_visuals/event_count_per_month.png'
        image_path = check_file_unique(image_path)
        plt.savefig(image_path)
        # Save the plot as an image
        plt.close()
        # I decided to use the matplotlib library to create various data visualizations
        print(f'\n-------------------------------------\nEvent count per month saved as {image_path}, download/view from the main menu.\n-------------------------------------')
    elif choice == '5':
        price = [extract_price(event['event_price']) for event in events if event.get('event_price', '').lower() not in ['sold out', 'free', 'donation']]
        # List comprehension to extract the prices from the events, not including 'sold out', 'free', and 'donation' events, then calling the extract_price function to extract the numeric part of the price
        plt.hist(price, bins=20, edgecolor='black')
        # A bin is a range of values that is used to group the data, the bins argument specifies the number of bins to use, the edgecolor argument specifies the color of the edges of the bars
        plt.title('Event Price Distribution')
        plt.xlabel('Price (£)')
        plt.ylabel('Frequency')
        image_path = 'data_visuals/event_price_distribution.png'
        image_path = check_file_unique(image_path)
        plt.savefig(image_path)
        plt.close()
        print(f'\n-------------------------------------\nEvent price distribution saved as {image_path}, download/view from the main menu.\n-------------------------------------')
    elif choice == '6':
        event_dates = [datetime.strptime(event['event_date_time'], '%Y-%m-%d %H:%M:%S').date() for event in events]
        # Grab all the event dates from the event_date_time field, parse the date and time to a datetime object, then extract the date
        date_counts = Counter(event_dates)
        # Count each occurrence of the different dates
        dates, counts = zip(*sorted(date_counts.items()))
        # Sort the dates and counts, then unpack them into two lists, it works by taking the date_counts Counter dictionary with items() which returns the key-value pairs as a list of tuples,
        # then sorted() sorts the tuples by the keys (dates), then zip() takes the sorted list of tuples and unpacks them into two seperate lists using the splat operator *
        plt.plot(dates, counts)
        # Create a line plot with the dates on the x-axis and the counts on the y-axis
        plt.title('Events Over Time')
        plt.xlabel('Date')
        plt.ylabel('Number of Events')
        plt.xticks(rotation=45)
        # Rotate the labels on the x-axis by 45 degrees for better readability
        plt.tight_layout()
        # Add padding to the plot
        image_path = 'data_visuals/event_dates_over_time.png'
        image_path = check_file_unique(image_path)
        plt.savefig(image_path)
        plt.close()
        print(f'\n-------------------------------------\nEvent dates over time saved as {image_path}, download/view from the main menu.\n-------------------------------------')
    elif choice == '7':
        print('Not implemented yet.')
    else:
        print('No valid comparison to display.')
     
      
def sort_events(events):
    if len(events) < 2:
        print('Not enough events to sort.')
        return
    
    print('\nWhat would you like to sort?')
    print('1. Free events')
    print('2. Cheapest events')
    print('3. Most expensive events')
    print('4. Events happening soon')
    print('5. Closest distance events')
    print('6. Main Menu')
    choice = input('Enter your choice: ').strip()
    
    # Lambda functions are a powerful tool for writing concise, one-off functions, especially useful in situations like sorting, filtering, and mapping.
    if choice == '1':
        free_events = [event for event in events if event.get('event_price', '').lower() in ['free', 'donation']]
    # Filter in events with event_price of 'free' and 'donation' with list comprehension, if not found, return an empty list
        display_events(free_events[::-1], 0, len(free_events), 'data-manipulation', 'None')
    # Display the sorted events from bottom to top
    elif choice == '2':
        cheap_events = [event for event in events if event.get('event_price', '').lower() not in ['sold out', 'free']]
    # Filter out events with event_price of 'free' and 'sold out' with list comprehension, if not found, return an empty list
        cheap_events_sorted = sorted(cheap_events, key=lambda x: extract_price(x.get('event_price', '0')))
    # Sort the remaining events in order based on event_price, the key argument specifies a custom sorting function for the sorted() method, and extracts a comparison key from each element.
    # The lambda function is given as the key argument to the sorted() method, it takes argument x that represents each element in the list, and extracts the price from the event_price field
    # with help from the extract_price function. The sorted() method will sort the events in ascending order based on the extracted price.
        display_events(cheap_events_sorted[::-1], 0, len(cheap_events_sorted), 'data-manipulation', 'None')
    # Display the sorted events from bottom to top, ::-1 is used to reverse the list
    elif choice == '3':
        paid_events = [event for event in events if event.get('event_price', '').lower() not in ['free', 'donation']]
    # Filter out events with event_price of 'free' and 'donation' with list comprehension, if not found, return an empty list
        expensive_events_sorted = sorted(paid_events, key=lambda x: extract_price(x.get('event_price', '0')), reverse=True)
    # Sort the remaining events in reverse order based on event_price, the key argument specifies a custom sorting function for the sorted() method, and extracts a comparison key from each element.
    # The lambda function is given as the key argument to the sorted() method, it takes argument x that represents each element in the list, and extracts the price from the event_price field
    # with help from the extract_price function. The sorted() method will sort the events in decending order based on the extracted price.
        display_events(expensive_events_sorted[::-1], 0, len(expensive_events_sorted), 'data-manipulation', 'None')
    # Display the sorted events from bottom to top , ::-1 is used to reverse the list
    elif choice == '4':
        current_time = datetime.now()
        soonest_events = sorted(
        [event for event in events if event.get('event_date_time') and datetime.strptime(event['event_date_time'], '%Y-%m-%d %H:%M:%S') > current_time],
        key=lambda x: datetime.strptime(x.get('event_date_time', ''), '%Y-%m-%d %H:%M:%S')
        )
        display_events(soonest_events[::-1], 0, len(soonest_events), 'data-manipulation', 'None')
    elif choice == '5':
        user_location = input('Enter your postcode or specific location: ')
        api_key = os.getenv('GOOGLE_MAPS_API_KEY')
        closest_events = find_closest_events(user_location, events, api_key)
        if closest_events:
            display_events(closest_events, 0, len(closest_events), 'data-manipulation', 'None')
        else:
            print('No events found or location cannot be geocoded.')
    elif choice == '6':
        print('Returning to the main menu.')
        main()
    

def event_manipulation_menu(events):
    while True:
        print('\nChoose an option to manipulate the events data:')
        print('1. Sort events')
        print('2. Compare events')
        print('3. Filter events')
        print('4. Main Menu')
        choice = input('Enter your choice: ').strip()
        
        if choice == '1':
            sort_events(events)
        elif choice == '2':
            compare_events(events)
        elif choice == '3':
            filter_events(events) # will be able to choose events based on collected tags
        elif choice == '4':
            print('Returning to the main menu.')
            main()

def search_events_in_collection():
    unique_search_keys = get_unique_search_keys()
    user_selection = 'data-manipulation'
    
    if len(unique_search_keys) == 0:
        print("No events found in the collection.")
        return
    
    print("-------------------------------------\nChoose a search key:")
    for i, key in enumerate(unique_search_keys, 1):
        print(f"{i}. {key}")
        
    choice = input("\nEnter the number of your choice: ").strip()
    all_events = list(collection.find({'search_key': unique_search_keys[int(choice) - 1]}))
    
    if not all_events:
        print("No events found for the selected search key.")
        return
    
    display_events(all_events, 0, len(all_events), user_selection, search_key=unique_search_keys[int(choice) - 1])

    save_choice = input('-------------------------------------\nWould you like to save the events to a CSV or Excel file? (C/E) or perform tasks on the data? (T): ').strip().lower()
    if save_choice == 'c':
        try:
            save_to_csv(all_events)
        except Exception as e:
            print(f"Error saving events to CSV: {e}")
    elif save_choice == 'e':
        try:
            save_to_excel(all_events)
        except Exception as e:
            print(f"Error saving events to Excel: {e}")
    elif save_choice == 't':
        event_manipulation_menu(all_events)


def view_all_events():
    all_events = list(collection.find({}))
    user_selection = 'data-manipulation'
    if len(all_events) == 0:
        print('No events found')
        return
    
    result = display_events(all_events, 0, len(all_events), user_selection, search_key='None')
    
    save_choice = input('-------------------------------------\nWould you like to save the events to a CSV or Excel file? (C/E) or perform tasks on the data? (T): ').strip().lower()
    if save_choice == 'c':
        try:
            save_to_csv(all_events)
        except Exception as e:
            print(f"Error saving events to CSV: {e}")
    elif save_choice == 'e':
        try:
            save_to_excel(all_events)
        except Exception as e:
            print(f"Error saving events to Excel: {e}")
    elif save_choice == 't':
        event_manipulation_menu(all_events)
        
    

def display_events(events, start_index, end_index, user_selection, search_key):
    collected_events = events[start_index:end_index]
    for data in collected_events:
        if isinstance(data, dict):
            show_date_time = data.get('show_date_time', 'No date and time available')
            summary = data['summary']
            truncated_summary = summary[:120] + '...' if len(summary) > 120 else summary
            if show_date_time != 'No date and time available':
                print(f'-------------------------------------\n{data["name"]},\n{data["location"]}\n{data["show_date_time"]}\nPrice: {data["event_price"]}\nSummary: {truncated_summary}\nURL: {data["url"]}')
            else:
                continue # Skip invalid event data
        else:
            continue # Skip invalid event data
            
    if user_selection == 'data-manipulation':
        return print('-------------------------------------\nEvents displayed in relevance bottom to top.')
    elif user_selection == 'eventbrite' or user_selection == 'eventbrite_top':
        save_to_mongodb(collection, search_key, collected_events)

    # Cache the events in the hashtable
    cache[search_key] = events
    
def search_events():
    product = input('Enter event type or name: ').replace(' ', '%20')
    location = input('Enter location: ').replace(' ', '%20')
    print('Would you like to enter a date? (Y/N)')
    date_choice = input('Enter your choice: ').strip().lower()
    
    start_date = ''
    end_date = ''
    day = ''
    
    if date_choice == 'y':
        print('Please enter an option: ')
        print('1. Today')
        print('2. Tomorrow')
        print('3. This weekend')
        print('4. Pick a date')
        day = input('Enter the number of choice: ')
        if day == '1':
            day = 'today'
        elif day == '2':
            day = 'tomorrow'
        elif day == '3':
            day = 'this-weekend'
        else:
            start_date = input('Enter the start date (YYYY-MM-DD): ')
            end_date = input('Enter the end date (YYYY-MM-DD): ')
        
    search_key = f'{product}_{location}'
    
    user_selection = 'eventbrite'

    spinner = Spinner("Fetching events...")
    spinner.start()

    unique_events = []
    page_number = 1

    try:
        # Check if the search term is in the cache
        if search_key in cache:
            spinner.stop()
            print("Using cached events from hashtable.")
            unique_events = cache[search_key]
        else:
            spinner.stop()
            spinner = Spinner("Scraping new events...")
            spinner.start()
            events_data, tags_counter = scrape_eventbrite_events(location, day, product, page_number, start_date, end_date)
            unique_events.extend(events_data)
            cache[search_key] = unique_events
    finally:
        spinner.stop()

    result = display_paginated_events(unique_events, search_key, user_selection, location, product, page_number)
    
    if result == 'new_search':
        main()
        return
    

def search_top_categories():
    categories = [
        "Home & Lifestyle", "Business", "Health", "Performing & Visual Arts",
        "Family & Education", "Holidays", "Music", "Community",
        "Hobbies", "Charity & Causes", "Food & Drink", "Science & Tech",
        "Sports & Fitness", "Travel & Outdoor", "Spirituality", "Nightlife",
        "Dating", "Film & Media", "Fashion", "Government", "Auto, Boat & Air",
        "School Activities"
    ]
    
    user_selection = 'eventbrite_top'

    def display_categories():
        print("\nPlease choose a category:")
        for i, category in enumerate(categories, 1):
            print(f"{i}. {category}")

    def get_user_choice():
        while True:
            user_input = input(f"Enter the number of your choice to find events in {location}: ")
            try:
                choice = int(user_input)
                if 1 <= choice <= len(categories):
                    return categories[choice - 1]
                # 1 is subtracted from the choice to get the correct index
                else:
                    print(f"Please enter a number between 1 and {len(categories)}.")
            except ValueError:
                print("Invalid input. Please enter a number.")

    def generate_slug(category):
        category = category.replace('&', 'and')
        return re.sub(r'\s+', '-', category.strip().lower())

    country = 'united-kingdom'
    location = input('Enter location: ').replace(' ', '')
    print('Would you like to enter a date? (Y/N)')
    date_choice = input('Enter your choice: ').strip().lower()
    
    start_date = ''
    end_date = ''
    day = ''
    
    if date_choice == 'y':
        print('Please enter an option: ')
        print('1. Today')
        print('2. Tomorrow')
        print('3. This weekend')
        print('4. Pick a date')
        day = input('Enter the number of choice: ')
        if day == '1':
            day = 'today'
        elif day == '2':
            day = 'tomorrow'
        elif day == '3':
            day = 'this-weekend'
        else:
            start_date = input('Enter the start date (YYYY-MM-DD): ')
            end_date = input('Enter the end date (YYYY-MM-DD): ')

    display_categories()
    category = get_user_choice()

    
    search_key = f'{generate_slug(category)}_{location}_{country}'
    spinner = Spinner("Fetching events...")
    spinner.start()

    unique_events = []
    page_number = 1

    try:
        # Check if the search term is in the cache
        if search_key in cache:
            spinner.stop()
            print("Using cached events from hashtable.")
            unique_events = cache[search_key]
        else:
            spinner.stop()
            spinner = Spinner("Scraping new events...")
            spinner.start()
            events_data, tags_counter, event_count = scrape_eventbrite_top_events(country, day, location, generate_slug(category), page_number, start_date, end_date)
            unique_events.extend(events_data)
            cache[search_key] = unique_events
    finally:
        spinner.stop()

    result = display_paginated_events(unique_events, search_key, user_selection, location, generate_slug(category), country, page_number)
        
    if result == 'new_search':
        main()
        return


def search_top_events():
    
    country = 'united-kingdom'
    location = input('Enter location: ').replace(' ', '')
    search_key = f'all_top_events_{location}_{country}'
    spinner = Spinner("Fetching events...")
    spinner.start()
    unique_events = []
    page_number = 1

    try:
        if search_key in cache:
            spinner.stop()
            print("Using cached events from hashtable.")
            unique_events = cache[search_key]
        else:
            spinner.stop()
            spinner = Spinner("Scraping new events...")
            spinner.start()
            events_data = scrape_eventbrite_top_events_no_category(location, country)
            unique_events.extend(events_data)
            cache[search_key] = unique_events
    finally:
        spinner.stop()

    result = display_paginated_events(unique_events, search_key, 'eventbrite_top', location, country, page_number)
    if result == 'new_search':
        main()
        return


def main():
    welcome = 0
    while True:
        if welcome < 1:
            print("-------------------------------------\nWelcome to Event Hoarder!\nSearch for events and they will be automatically be saved to a database so you can\nperform sorting, comparing or filtering tasks, also print to CSV\n-------------------------------------")
            welcome += 1
        print("\nChoose an option:")
        print("1. Quick Search & Collect")
        print("2. Search & Collect Top Events")
        print("3. Search & Collect Top Categories")
        print("4. View Collected Events")
        print("5. Download saved Excel , CSV or data visuals")
        print("6. Exit")
        print("#. Clear Database")
        choice = input("Enter your choice: ").strip()

        if choice == '1':
            search_events()
        elif choice == '2':
            search_top_events()
        elif choice == '3':
            search_top_categories()
        elif choice == '4':
            collection_menu()
        elif choice =='5':
            start_flask_server()    
        elif choice == '6':
            print("-------------------------------------\nExiting the program\n-------------------------------------.")
            sys.exit()
        elif choice == '#':
            collection.delete_many({})
            print('-------------------------------------\nDatabase cleared\n-------------------------------------.')
            main()
        else:
            print("\nInvalid choice. Please try again.")


def display_paginated_events(unique_events, search_key, user_selection, location=None, country=None, category_slug=None, product=None, page_number=1):
    page_size = 5
    total_events = len(unique_events)
    current_page = 0
    tags_counter = Counter()

    while current_page * page_size < total_events:
        start_index = current_page * page_size
        end_index = min(start_index + page_size, total_events)
        display_events(unique_events, start_index, end_index, user_selection, search_key)
        
        if end_index >= total_events:
            # Fetch more events if available
            page_number += 1
            spinner = Spinner("Fetching more events...")
            spinner.start()
            try:
                if user_selection == 'eventbrite':
                    events_data, new_tags_counter = scrape_eventbrite_events(location, product, page_number)
                elif user_selection == 'eventbrite_top':
                    events_data, new_tags_counter, event_count = scrape_eventbrite_top_events(country, location, category_slug, page_number)
                else:
                    break  # No more events to fetch

                if not events_data:
                    break  # No more events to fetch

                unique_events.extend(events_data)
                tags_counter.update(new_tags_counter)
                total_events = len(unique_events)
            finally:
                spinner.stop()

        user_input = input("-------------------------------------\nPress 'Y' to see more events, 'S' to start a new search, or any other key to exit: ").strip().lower()
        if user_input == 's':
            return 'new_search'
        elif user_input == 'y':
            current_page += 1
        else:
            print("Exiting the program.")
            sys.exit()

    spinner = Spinner("Loading most common tags...")
    spinner.start()
    most_common_tags = tags_counter.most_common(6)
    spinner.stop()

    print(f'\nThe most common tags are:')
    for tag, count in most_common_tags:
        print(f'{tag}: {count}')
        
    return 'done'
        
if __name__ == "__main__":
    main()